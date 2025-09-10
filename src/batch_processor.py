"""
Batch Processor for ROI Calculator
Handles bulk processing of multiple scenarios and comparisons
"""

import pandas as pd
import numpy as np
import json
import os
import tempfile
from datetime import datetime
from typing import Dict, List, Optional, Any, Union
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import traceback

from roi_calculator import ROICalculator
from template_manager import TemplateManager


class BatchProcessor:
    """Processes multiple ROI calculations in parallel and generates comparison reports"""
    
    def __init__(self, max_workers: int = 4):
        self.max_workers = max_workers
        self.roi_calculator = ROICalculator()
        self.template_manager = TemplateManager()
        self.results_cache = {}
    
    def import_from_csv(self, file_path: str, mapping: Dict[str, str] = None) -> List[Dict]:
        """Import scenarios from CSV file"""
        try:
            # Read CSV file
            df = pd.read_csv(file_path)
            
            # Default column mapping
            default_mapping = {
                'company_name': 'company_name',
                'annual_revenue': 'annual_revenue',
                'monthly_orders': 'monthly_orders',
                'avg_order_value': 'avg_order_value',
                'labor_costs': 'labor_costs',
                'shipping_costs': 'shipping_costs',
                'error_costs': 'error_costs',
                'inventory_costs': 'inventory_costs',
                'service_investment': 'service_investment'
            }
            
            # Use provided mapping or default
            column_mapping = mapping or default_mapping
            
            # Validate required columns exist
            required_cols = [
                'annual_revenue', 'monthly_orders', 'avg_order_value',
                'labor_costs', 'shipping_costs', 'error_costs',
                'inventory_costs', 'service_investment'
            ]
            
            missing_cols = []
            for req_col in required_cols:
                mapped_col = column_mapping.get(req_col, req_col)
                if mapped_col not in df.columns:
                    missing_cols.append(mapped_col)
            
            if missing_cols:
                raise ValueError(f"Missing required columns: {', '.join(missing_cols)}")
            
            # Convert dataframe to scenarios list
            scenarios = []
            for index, row in df.iterrows():
                scenario = {}
                
                # Map columns
                for target_col, source_col in column_mapping.items():
                    if source_col in df.columns:
                        value = row[source_col]
                        
                        # Handle numeric conversion
                        if target_col != 'company_name' and pd.notna(value):
                            try:
                                scenario[target_col] = float(value)
                            except (ValueError, TypeError):
                                scenario[target_col] = 0
                        else:
                            scenario[target_col] = str(value) if pd.notna(value) else f"Scenario_{index + 1}"
                
                # Add scenario identifier
                if 'company_name' not in scenario or pd.isna(scenario['company_name']):
                    scenario['company_name'] = f"Scenario_{index + 1}"
                
                scenarios.append(scenario)
            
            return scenarios
            
        except Exception as e:
            raise Exception(f"Error importing CSV: {str(e)}")
    
    def import_from_excel(self, file_path: str, sheet_name: str = None, 
                         mapping: Dict[str, str] = None) -> List[Dict]:
        """Import scenarios from Excel file"""
        try:
            # Read Excel file
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            # Convert to CSV format and use CSV import logic
            temp_csv = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False)
            df.to_csv(temp_csv.name, index=False)
            temp_csv.close()
            
            try:
                scenarios = self.import_from_csv(temp_csv.name, mapping)
                return scenarios
            finally:
                # Clean up temp file
                os.unlink(temp_csv.name)
                
        except Exception as e:
            raise Exception(f"Error importing Excel: {str(e)}")
    
    def process_scenarios_parallel(self, scenarios: List[Dict], 
                                 progress_callback: Optional[callable] = None) -> List[Dict]:
        """Process multiple scenarios in parallel"""
        try:
            results = []
            total_scenarios = len(scenarios)
            
            def calculate_single_scenario(scenario_data):
                """Calculate ROI for a single scenario"""
                try:
                    scenario_id = scenario_data.get('company_name', 'Unknown')
                    calculator = ROICalculator()
                    roi_results = calculator.calculate_roi(scenario_data)
                    
                    return {
                        'scenario_id': scenario_id,
                        'inputs': scenario_data,
                        'results': roi_results,
                        'success': True,
                        'error': None
                    }
                except Exception as e:
                    return {
                        'scenario_id': scenario_data.get('company_name', 'Unknown'),
                        'inputs': scenario_data,
                        'results': None,
                        'success': False,
                        'error': str(e)
                    }
            
            # Use ThreadPoolExecutor for I/O bound tasks
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                # Submit all scenarios
                future_to_scenario = {
                    executor.submit(calculate_single_scenario, scenario): i 
                    for i, scenario in enumerate(scenarios)
                }
                
                # Process completed futures
                completed = 0
                for future in as_completed(future_to_scenario):
                    result = future.result()
                    results.append(result)
                    
                    completed += 1
                    if progress_callback:
                        progress_callback(completed, total_scenarios)
            
            # Sort results by original order
            results.sort(key=lambda x: scenarios.index(next(s for s in scenarios 
                                                           if s.get('company_name') == x['scenario_id'])))
            
            return results
            
        except Exception as e:
            raise Exception(f"Error processing scenarios: {str(e)}")
    
    def generate_comparison_matrix(self, results: List[Dict]) -> Dict:
        """Generate comprehensive comparison matrix from batch results"""
        try:
            if not results:
                raise ValueError("No results to compare")
            
            # Filter successful results
            successful_results = [r for r in results if r['success']]
            
            if not successful_results:
                raise ValueError("No successful calculations to compare")
            
            # Extract key metrics for comparison
            comparison_data = []
            
            for result in successful_results:
                roi_metrics = result['results'].get('roi_metrics', {})
                financial_metrics = result['results'].get('financial_metrics', {})
                projections = result['results'].get('projections', {})
                
                row = {
                    'scenario_id': result['scenario_id'],
                    'annual_revenue': result['inputs'].get('annual_revenue', 0),
                    'service_investment': result['inputs'].get('service_investment', 0),
                    'roi_year_1': roi_metrics.get('first_year_roi', 0),
                    'roi_year_3': projections.get('year_3', {}).get('roi_percentage', 0),
                    'payback_months': roi_metrics.get('payback_period_months', 0),
                    'annual_savings': roi_metrics.get('annual_savings', 0),
                    'monthly_savings': roi_metrics.get('monthly_savings', 0),
                    'npv': financial_metrics.get('npv', 0),
                    'irr': financial_metrics.get('irr', 0) * 100,  # Convert to percentage
                    'total_3_year_benefit': projections.get('year_3', {}).get('cumulative_savings', 0)
                }
                comparison_data.append(row)
            
            # Create DataFrame for analysis
            df = pd.DataFrame(comparison_data)
            
            # Calculate rankings and statistics
            metrics_to_rank = ['roi_year_1', 'roi_year_3', 'annual_savings', 'npv', 'irr', 'total_3_year_benefit']
            inverse_metrics = ['payback_months', 'service_investment']  # Lower is better
            
            # Add rankings
            for metric in metrics_to_rank:
                df[f'{metric}_rank'] = df[metric].rank(ascending=False, method='min')
            
            for metric in inverse_metrics:
                df[f'{metric}_rank'] = df[metric].rank(ascending=True, method='min')
            
            # Calculate percentile rankings
            for metric in metrics_to_rank:
                df[f'{metric}_percentile'] = df[metric].rank(pct=True) * 100
            
            # Statistical analysis
            stats = {}
            for metric in metrics_to_rank + inverse_metrics:
                if metric in df.columns:
                    stats[metric] = {
                        'mean': float(df[metric].mean()),
                        'median': float(df[metric].median()),
                        'std': float(df[metric].std()),
                        'min': float(df[metric].min()),
                        'max': float(df[metric].max()),
                        'q25': float(df[metric].quantile(0.25)),
                        'q75': float(df[metric].quantile(0.75))
                    }
            
            # Find top performers
            top_performers = {
                'highest_roi_year_1': df.loc[df['roi_year_1'].idxmax()].to_dict(),
                'highest_roi_year_3': df.loc[df['roi_year_3'].idxmax()].to_dict(),
                'fastest_payback': df.loc[df['payback_months'].idxmin()].to_dict(),
                'highest_npv': df.loc[df['npv'].idxmax()].to_dict(),
                'best_overall': df.loc[df[['roi_year_1_rank', 'roi_year_3_rank', 'npv_rank']].mean(axis=1).idxmin()].to_dict()
            }
            
            # Risk analysis
            risk_analysis = self._analyze_batch_risk(df)
            
            return {
                'total_scenarios': len(results),
                'successful_scenarios': len(successful_results),
                'failed_scenarios': len(results) - len(successful_results),
                'comparison_matrix': df.to_dict('records'),
                'statistical_summary': stats,
                'top_performers': top_performers,
                'risk_analysis': risk_analysis,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            raise Exception(f"Error generating comparison matrix: {str(e)}")
    
    def _analyze_batch_risk(self, df: pd.DataFrame) -> Dict:
        """Analyze risk across batch of scenarios"""
        try:
            risk_metrics = {}
            
            # ROI volatility
            roi_std = df['roi_year_1'].std()
            roi_mean = df['roi_year_1'].mean()
            roi_cv = (roi_std / roi_mean) if roi_mean != 0 else 0  # Coefficient of variation
            
            # Payback period risk
            payback_risk = (df['payback_months'] > 24).sum() / len(df) * 100  # % with >2 year payback
            
            # Investment efficiency spread
            inv_efficiency = df['annual_savings'] / df['service_investment']
            efficiency_spread = inv_efficiency.max() - inv_efficiency.min()
            
            # Scenario reliability (% of successful vs failed scenarios)
            reliability = len(df) / (len(df) + 0)  # Assuming successful results only in df
            
            risk_metrics = {
                'roi_volatility': {
                    'coefficient_of_variation': float(roi_cv),
                    'risk_level': 'High' if roi_cv > 0.5 else 'Medium' if roi_cv > 0.2 else 'Low'
                },
                'payback_risk': {
                    'long_payback_percentage': float(payback_risk),
                    'risk_level': 'High' if payback_risk > 50 else 'Medium' if payback_risk > 25 else 'Low'
                },
                'investment_efficiency': {
                    'efficiency_spread': float(efficiency_spread),
                    'min_efficiency': float(inv_efficiency.min()),
                    'max_efficiency': float(inv_efficiency.max()),
                    'median_efficiency': float(inv_efficiency.median())
                },
                'overall_risk_score': self._calculate_risk_score(roi_cv, payback_risk, efficiency_spread)
            }
            
            return risk_metrics
            
        except Exception as e:
            return {'error': f"Risk analysis failed: {str(e)}"}
    
    def _calculate_risk_score(self, roi_cv: float, payback_risk: float, efficiency_spread: float) -> str:
        """Calculate overall risk score"""
        score = 0
        
        # ROI volatility score (0-40 points)
        if roi_cv > 0.5:
            score += 40
        elif roi_cv > 0.3:
            score += 25
        elif roi_cv > 0.1:
            score += 10
        
        # Payback risk score (0-30 points)
        if payback_risk > 50:
            score += 30
        elif payback_risk > 25:
            score += 20
        elif payback_risk > 10:
            score += 10
        
        # Efficiency spread score (0-30 points)
        if efficiency_spread > 2:
            score += 30
        elif efficiency_spread > 1:
            score += 20
        elif efficiency_spread > 0.5:
            score += 10
        
        if score >= 70:
            return 'High Risk'
        elif score >= 40:
            return 'Medium Risk'
        else:
            return 'Low Risk'
    
    def export_to_excel(self, comparison_data: Dict, file_path: str = None) -> str:
        """Export batch results to Excel with multiple sheets"""
        try:
            if not file_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                file_path = f"roi_batch_analysis_{timestamp}.xlsx"
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # 1. Summary Sheet
            ws_summary = wb.create_sheet("Summary")
            self._create_summary_sheet(ws_summary, comparison_data)
            
            # 2. Comparison Matrix Sheet
            ws_matrix = wb.create_sheet("Comparison Matrix")
            self._create_comparison_sheet(ws_matrix, comparison_data)
            
            # 3. Statistical Analysis Sheet
            ws_stats = wb.create_sheet("Statistical Analysis")
            self._create_statistics_sheet(ws_stats, comparison_data)
            
            # 4. Top Performers Sheet
            ws_top = wb.create_sheet("Top Performers")
            self._create_top_performers_sheet(ws_top, comparison_data)
            
            # 5. Risk Analysis Sheet
            ws_risk = wb.create_sheet("Risk Analysis")
            self._create_risk_analysis_sheet(ws_risk, comparison_data)
            
            # Save workbook
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            raise Exception(f"Error exporting to Excel: {str(e)}")
    
    def _create_summary_sheet(self, worksheet, data: Dict):
        """Create summary sheet in Excel"""
        # Title
        worksheet['A1'] = "ROI Batch Analysis Summary"
        worksheet['A1'].font = Font(size=16, bold=True)
        
        # Basic statistics
        row = 3
        worksheet[f'A{row}'] = "Analysis Overview"
        worksheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        worksheet[f'A{row}'] = "Total Scenarios:"
        worksheet[f'B{row}'] = data.get('total_scenarios', 0)
        row += 1
        
        worksheet[f'A{row}'] = "Successful Calculations:"
        worksheet[f'B{row}'] = data.get('successful_scenarios', 0)
        row += 1
        
        worksheet[f'A{row}'] = "Failed Calculations:"
        worksheet[f'B{row}'] = data.get('failed_scenarios', 0)
        row += 1
        
        worksheet[f'A{row}'] = "Analysis Generated:"
        worksheet[f'B{row}'] = data.get('generated_at', '')
        row += 2
        
        # Top performers summary
        if 'top_performers' in data:
            worksheet[f'A{row}'] = "Top Performers"
            worksheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            top_performers = data['top_performers']
            for category, performer in top_performers.items():
                worksheet[f'A{row}'] = category.replace('_', ' ').title() + ":"
                worksheet[f'B{row}'] = performer.get('scenario_id', 'Unknown')
                row += 1
    
    def _create_comparison_sheet(self, worksheet, data: Dict):
        """Create comparison matrix sheet"""
        if 'comparison_matrix' not in data:
            return
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(data['comparison_matrix'])
        
        # Add headers
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    def _create_statistics_sheet(self, worksheet, data: Dict):
        """Create statistical analysis sheet"""
        if 'statistical_summary' not in data:
            return
        
        worksheet['A1'] = "Statistical Summary"
        worksheet['A1'].font = Font(size=16, bold=True)
        
        row = 3
        stats = data['statistical_summary']
        
        # Headers
        headers = ['Metric', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Q25', 'Q75']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        
        row += 1
        
        # Data
        for metric, values in stats.items():
            worksheet.cell(row=row, column=1, value=metric)
            worksheet.cell(row=row, column=2, value=values.get('mean', 0))
            worksheet.cell(row=row, column=3, value=values.get('median', 0))
            worksheet.cell(row=row, column=4, value=values.get('std', 0))
            worksheet.cell(row=row, column=5, value=values.get('min', 0))
            worksheet.cell(row=row, column=6, value=values.get('max', 0))
            worksheet.cell(row=row, column=7, value=values.get('q25', 0))
            worksheet.cell(row=row, column=8, value=values.get('q75', 0))
            row += 1
    
    def _create_top_performers_sheet(self, worksheet, data: Dict):
        """Create top performers sheet"""
        if 'top_performers' not in data:
            return
        
        worksheet['A1'] = "Top Performing Scenarios"
        worksheet['A1'].font = Font(size=16, bold=True)
        
        row = 3
        top_performers = data['top_performers']
        
        for category, performer in top_performers.items():
            worksheet[f'A{row}'] = category.replace('_', ' ').title()
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            worksheet[f'A{row}'] = "Scenario:"
            worksheet[f'B{row}'] = performer.get('scenario_id', 'Unknown')
            row += 1
            
            # Add key metrics
            key_metrics = ['roi_year_1', 'roi_year_3', 'payback_months', 'annual_savings', 'npv']
            for metric in key_metrics:
                if metric in performer:
                    worksheet[f'A{row}'] = metric.replace('_', ' ').title() + ":"
                    worksheet[f'B{row}'] = performer[metric]
                    row += 1
            
            row += 1  # Add spacing
    
    def _create_risk_analysis_sheet(self, worksheet, data: Dict):
        """Create risk analysis sheet"""
        if 'risk_analysis' not in data:
            return
        
        worksheet['A1'] = "Risk Analysis"
        worksheet['A1'].font = Font(size=16, bold=True)
        
        row = 3
        risk_data = data['risk_analysis']
        
        for category, risk_info in risk_data.items():
            if isinstance(risk_info, dict):
                worksheet[f'A{row}'] = category.replace('_', ' ').title()
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for key, value in risk_info.items():
                    worksheet[f'A{row}'] = "  " + key.replace('_', ' ').title() + ":"
                    worksheet[f'B{row}'] = value
                    row += 1
                
                row += 1  # Add spacing
    
    def save_batch_results(self, results: List[Dict], filename: str = None) -> str:
        """Save batch results to JSON file"""
        try:
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"roi_batch_results_{timestamp}.json"
            
            # Prepare data for JSON serialization
            serializable_results = []
            for result in results:
                serializable_result = {
                    'scenario_id': result['scenario_id'],
                    'success': result['success'],
                    'error': result['error'],
                    'inputs': result['inputs']
                }
                
                # Handle results serialization
                if result['results']:
                    # Convert numpy types to native Python types
                    serializable_result['results'] = self._convert_for_json(result['results'])
                
                serializable_results.append(serializable_result)
            
            # Save to file
            with open(filename, 'w') as f:
                json.dump({
                    'batch_results': serializable_results,
                    'generated_at': datetime.now().isoformat(),
                    'total_scenarios': len(results),
                    'successful_scenarios': len([r for r in results if r['success']])
                }, f, indent=2)
            
            return filename
            
        except Exception as e:
            raise Exception(f"Error saving batch results: {str(e)}")
    
    def _convert_for_json(self, obj: Any) -> Any:
        """Convert numpy types and other non-serializable types to JSON-compatible types"""
        if isinstance(obj, dict):
            return {k: self._convert_for_json(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._convert_for_json(item) for item in obj]
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        else:
            return obj
    
    def load_batch_results(self, filename: str) -> Dict:
        """Load previously saved batch results"""
        try:
            with open(filename, 'r') as f:
                data = json.load(f)
            
            return data
            
        except Exception as e:
            raise Exception(f"Error loading batch results: {str(e)}")